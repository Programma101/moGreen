#-*- coding: utf-8 -*-
#nella versione 28e ho spostato save store da interno alla funzione scrivo dati giornalieri in altri fogli
#a fuori visto che l'importazione di opepyxle e il caricamento dei fogli è esterno dopo che sono stati aggiornati
#tutti e 3 i fogli con i dati giornalieri.
#nella versione 28d ho eliminatolo del tutto il controllo su gSheet del limite accensione lampada inserendo un valore fisso a 700 lux
#nel tentativo di fare eseguire le foto anche se non si è connessi. 
#nella cersione 28c ho ripristinato le operazioni di mezzanotte che partono da 00 minuti e finiscono a 7
#nella versione 28b provo a fare funzionare sriviDatiGiornalieriInAltriFogli
#nella versione 28a ho cercato di provare per le operazioni che salta alla mezzanotte di togliere min 01 e 02 e arrivare fino a 09
#nella versione 28 ho ripristinato foglio48 ore
#Nella versione 27 ho esternalizzato tutte le letture dei sensori e la scrittura di questi in store.
#Elimino anche i controllo valori rilevati e il riempimento delle righe vuote
#nella versione 26a ho scritto i print di di tutte le azioni anche se non vengono eseguite a quell'ora
#ho sostituito tutti i 4 min con 8 minuti 
#nella versione 26 ho cercato di correggere gli script di lettura del sensore VEML7700 e BME280 per non creare
#conflitto d'inidirizzi I2C (chiusura smbus dopo la lettura ed intervallo di 1 secondo tra una lettura ed un'altra)
#nella versione 25 ho introdotto il salvataggio di store giornaliero nella memoria usb
# vella versione 24 del 7 aprile introduco anche la funzione
#doOraE8min in quanto i dati giornalieri non vengono aggiornati su gSheet
#nella versione 23 in caso d'interruzione corrente elettrica vengono iserite righe anche nel foglio ultime48ore
#e nel foglio datiOrari che da questa versione in poi ho aggiunto a store e gSheet
#nella versione 22 ho inserito la funzione do8min che userò per le azioni da attuare ogni 30 min
#nella versione 21 inserisco 2 chunk: 1) in caso d'interruzione di corrente al ripristino
#inserisce le righe mancanti in store con data corretta ma ripetendo sempre gli stessi valori dell'ultimo rilievo
#2)il secondo chunk verifica se i valori dei rilievi sono testuali invece che numerici e se sono fuori da un range
#prestabilito


#nella versione 20 cercherò di distinguere meglio gli errori inserendo nel nome del file
#il punto esatto dove è capitato, per il rilievo dei sensori l'ho già fatto. 

#Nella versione 16 si cerca di suddividere il locale da quello che va su Internet in modo che in caso di disconnessione non ci siano errori

#******* V e r s i o n e   c o n   scrivi48oreStore e scrivi dati giornalieri in altri fogli  funzionanti  *********
#******* Versione senza il rilievo sensori BS perchÉ perdono l'indirizzo, in attesa di sostituzione e di capire
#******* La versione 18bsoloBSAddr24 è quella di prova una volta abbassato il clock del I2C da 100KHz a 50 
#******* La versione 19 è del tutto simile alla 18f ma lo script del sensore luce è stato verificato dal costruttore
#******* per poter rilevare fino a 120000 lux bisogna usare una frequenza di 25ms e pertanto l'errore può
#******* essere anche del 20 percento. sarebbe meglio una procedura di calibrazione che non conosco
#******* nel presente script di VEML7700 c'è la formula di correzione per rilievi maggiori di 100 lx
#******* In questa versione c'è anche aggiunta righe in caso di interruzione corrente  
#******* in seguito a test il presente potrebbe essere quello definitivo salvo correzioni input
#******* che non sono essenziali prima del test in campo, se c'è tempo le aggiuNngo


#verifica se il precedente dato è stato 15 minuti prima
"""
Se la corrente elettrica va via per parecchio tempo, tipo ore, vorrei che la tabella
venisse compilata saltando le righe per quanti sono i rilievi mancanti.
"""
# A T T E N Z I O N E ! Questo chunk va attaccato prima del rilievo dei sensori,
#in pratica subito dopo le definizioni di googleSheet
import traceback


#******* M O D U L I   E S C L U S O   Q U E L L I   D E I   S E N S O R I *********

import datetime
from datetime import datetime
#import RPi.GPIO as GPIO
#GPIO.setwarnings(False)
import time
import sys

#import oauth2client.client
#import json
import os
import ftplib


#******* F U N Z I O N I     E S C L U S E   Q U E L L E   D E I   S E N S O R I *********
"""

#***************** e s e g u e   u n a  f u n z i o n e   in    d e t e r m i n a t i   g i o r n i   o r a  e  m i n u t i  *****************
#def stampa():
#    print("sto eseguendo la funzione")        
def doGiorniOraEMin(giornoA, giornoB, oraA, minA, minB, minC,minD, funzione,*arg):
    import datetime    
    now=datetime.datetime.now()# chiedo che ora è
    print (now)
    print (type(now))
    minuti=now.strftime('%M')#chiedo in questo momento in che minuti dell'ora siamo
    ora=now.strftime('%H')#chiedo in questo momento in che ora siamo
    giorno=now.strftime('%A')
    print (ora)
    print (minuti)
    print(giorno)
    if (giorno==giornoA and ora==oraA and (minuti==minA or minuti==minB or minuti==minC or minuti==minD)) or (giorno==giornoB and ora==oraA and (minuti==minA or minuti==minB or minuti==minC or minuti==minD)):
      
        print("Visto che siamo al giorno all'ora e minuti giusti fai una qualche azione")
        funzione(*arg)
    else:
        print("Non siamo ancora al giorno, all'ora e minuti giusti non eseguo la funzione")
# A T T E N Z I O N E ! ! !  Nel chiamare la funzione  scrivere i giorni della settimana in inglese con la prima lettera maiuscola
#doGiorniOraEMin("Monday","Tuesday","23","06","07","09","08",stampa)

"""





def ultimaConValori():
    import openpyxl
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx')
    ws=wb["Sheet1"]
    #ws1=wb["ultime48ore"]
    #ws2=wb["ultimaSettimana"]
    #ws3=wb["ultime2settimane"]
    #ws4=wb["ultimoMese"]
    #ws5=wb["datiGiornalieri"]
    #ws8=wb["datiOrari"]
    from time import sleep
    from picamera import PiCamera #la parte di picamera relativa alle foto, penso
    #Devo inserire le librerie che sevono per leggere i dati da store invece
    #da googleFogli
    #introduco le librerie e il codice per leggere i dati da stor    
    #codice di picamera
    camera=PiCamera()
    camera.resolution = (2592, 1944)
    camera.annotate_text_size=65
    #camera.annotate_foreground=0, 0, 0
    #camera.annotate_background=Color('255, 255, 255')al momento così le pecette non funzionano provo con black
    #camera.annotate_background = picamera.Color('black') neanxche così funziona
    camera.annotate_background = True#FUNZIONA!!!
    #apro la cartella store
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx')
    #e definisco il foglio ws
    #ws=wb.get_sheet_by_name('Sheet1'), non funziona: ho usato wb.active
    ws=wb["Sheet1"]
    #conto le righe e il numero di righe lo chiamo r
    r=ws.max_row
    #ATTENZIONE TUTTI I DATI DEVONO ESSERE STRINGHE PER POTER ESSERE STAMPATI SULLA FOTO 
    #estraggo da store dataOra
    dataOra=ws.cell(row=r,column=1).value
    dataOra=str(dataOra)
    #Estraggo da store la temperatura dell'aria
    TAria=ws.cell(row=r,column=2).value
    #print TAria #per prova
    TAria=str(TAria)
    #print (TAria)
    #Estraggo da store UR
    UR=ws.cell(row=r,column=3).value
    #print (UR) #per prova
    UR=str(UR)
    #print type(UR)
    #Estraggo da store BS
    Lux=ws.cell(row=r,column=4).value
    #print BS #per prova
    Lux=str(Lux)
    #print type(BS)
    #Estraggo da store LA
    BS1=ws.cell(row=r,column=5).value
    #print LA #per prova
    BS1=str(BS1)
    #print type(LA)
    #Estraggo da store LP
    TS1=ws.cell(row=r,column=6).value
    #print LP #per prova
    TS1=str(TS1)
    #print type(LP)
    #Estraggo da store 
    BS2=ws.cell(row=r,column=7).value
    #print TS #per prova
    BS2=str(BS2)
    #print type(TS)
    #Estraggo da store LP
    TS2=ws.cell(row=r,column=8).value
    #print LP #per prova
    TS2=str(TS2)
    #print type(LP)
    #provo a mettere tutti i valori in un unico testo
    testo=dataOra+", TA "+TAria+", UR "+UR+", Lux "+Lux+", BS1 "+BS1+", TS1 "+TS1+", BS2 "+BS2+", TS2 "+TS2
    print("inserisco i valori sulla foto ",testo)
    camera.start_preview()
    camera.annotate_text=testo
    sleep(2)
    camera.capture('/home/pi/plant+out/PlantToWeb/ultima.jpg') #la inserisce in PlantToWeb e la chiama ultima.jpg
    camera.stop_preview()
    camera.close()
    del openpyxl
    print("Ho scattato la foto 'utlima.jpg' che si trova in '/home/pi/plant+out/PlantToWeb'")
    


def fotoCondizioneLux():
    #per prima cosa devo verificare se c'è connessione
    print("recupero il dato lux da store")
    #importo la libreria per leggere su excell: 
    import openpyxl
    wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
    ws=wb["Sheet1"]
    #conta il numero di righe già scritte per scrivere i dati nella riga giusta
    nr=ws.max_row
    print ("la riga di lettura di ws di store è ",nr)
    lux=ws['D'+str(nr)].value
    print("l'ultimo rilievo della luce è ",lux)
    print("deimporto la libreria openpyxl")
    del openpyxl
    print("la luce è ",lux)
    LimAccLampada=700
    print("la soglia per l'accenzione lampada è ",LimAccLampada)
    if lux >LimAccLampada: #Se vuole che la lampada si accenda sempre mettere un valore >10000
        ultimaConValori()
        print("ho fatto la foto alla pianta senza accendere la lampada perche' c'è sufficiente luce")
    else:
        import time
        import RPi.GPIO as GPIO
        GPIO.setwarnings(False) 
        GPIO.setmode(GPIO.BCM) #imposta il tipo di numerazione dei GPio in base al processore
        GPIO.setup(6,GPIO.OUT) #imposta il la porta GPIO 5 (secondo la numerazione BMC) come dati in uscita
        GPIO.output(6, False) #mette il GPIO 5 a 0.0v. la lampada si accende
        print("ho acceso la lampada perché la luce non è sufficiente per la foto")
        time.sleep(4)
        ultimaConValori()
        print("ho fatto la foto alla pianta accendendo la lampada")
        time.sleep(1)
        GPIO.output(6, True) #mette il pin a 3.3 v.  cioè tiene la lampada spenta
        GPIO.cleanup() #per assicurare l'uscita dal modulo GPIO
        print("ho spento la lampada")
        
def invioFotoSuWeb():
    #ho inserito la verifica della connessione su do
    print("...... attendi qualche secondo che sto inviando ultima.jpg su web")
    #una volta fatta la foto la lancio su web
    import ftplib #importiamo la libreria che useremo per gestire la connessione FTP
    ftp = ftplib.FTP('rpiplant.altervista.org','rpiplant','aPQq7jNJpWzS') # Si connette
    fp = open('/home/pi/plant+out/PlantToWeb/ultima.jpg','rb') # Imposta il file da inviare, apriamo uno stream per il file
    #di default siamo nella cartella root del sito / - se vogliamo spostarci in un'altra directory è sufficiente scrivere: ftp.cwd('directory')
    ftp.storbinary('STOR ultima.jpg', fp) # Invia il file 
    fp.close() # Chiude lo stream del file
    ftp.quit() # Chiude la connessione
    print("ultima.jpg inviata a http://rpiplant.altervista.org/ultima.jpg")
 
    
    
def copiaUltimaInFotoDiPiante():
    #da qui in poi sposto la foto Ultima in foto di piante con nome foto*****
    print ("sposto la foto Ultima in fotoDiPiante attribuendo un numero progressivo alla foto")
    import os
    import shutil
    #chiampo path l'indirizzo assoluto della cartella
    path="/media/pi/45D5-C1E61/FotoDiPiante"
    #path="/home/pi/plant+out/Prove"
    #os.chdir(path)
    #path="/home/pi/plant+out/PlantToWeb" #perle prove
    if os.listdir(path)==[]:
        print("la cartella è vuota")#che dopo aver fatto il video la cartella FotoDiPiante rimane vuota
        nomeFoto="Foto00001.jpg"
        vecchioNome = "/home/pi/plant+out/PlantToWeb/ultima.jpg"
        nuovoNome = "/home/pi/plant+out/PlantToWeb/"+nomeFoto
        path1="/home/pi/plant+out/PlantToWeb/"+nomeFoto
        path2="/media/pi/45D5-C1E61/FotoDiPiante"
        print(nuovoNome)
        os.rename(vecchioNome,nuovoNome)
        shutil.move(path1,path2)
        print("ho spostato ultima.jpg in FotoDiPiante e l'ho chiamato Foto00001.jpg")
    else:
        print("la cartella non è vuota, segue la lista dei files contenuti nella cartella")
        path="/media/pi/45D5-C1E61/FotoDiPiante"
        #siccome il file in python non è posizionato nella directory su cui agisce
        #devo cambiare la directory di lavoro corrente (cwd) altrimenti
        #il comando os.listdir non funziona
        os.chdir(path)#con questo comando cambio directory di lavoro
        #ed infatti se chiedo la cwd mi dà quella di fotoDiPiante
        #non quella dov'è posizionato il file
        cwd = os.getcwd()  # Get the current working directory (cwd)
        files = os.listdir(cwd)  # Get all the files in that directory
        print("Files in %r: %s" % (cwd, files))
        #segue finalmente il comando che estrae il nome dell'ultima foto
        ultimoFile=max(os.listdir(path), key=os.path.getctime)
        print("il nome dell'ultimo file nella directory fotoDiPiante e' ",ultimoFile)
        #print(type(ultimoFile))
        #Ora dovrei estrarre dal nome file che è una stringa dalla 5a alla 9a lettera
        nrFoto=ultimoFile[4:9]
        print(nrFoto)
        print(type(nrFoto))
        nrFoto=int(nrFoto)# trasformo la stringa in numero per poter addizionare 1
        print(type(nrFoto))     
        print(nrFoto, "ora è trasformato in numero intero")
        nrFoto=nrFoto+1
        print("ho aggiunto 1 al numero della foto che è diventato",nrFoto)
   
        if nrFoto<10:
            nrFoto=str(nrFoto) #per concatenarlo ad altre stringhe lo trasformo in stringa
            nomeFoto="Foto0000"+nrFoto+"."+"jpg"
            nrFoto=int(nrFoto) #devo ritrasformarlo in numero per poter capire se è < o > a qualche altro numero
            #per potere eseguire la condizione che segue.
        if nrFoto>9 and nrFoto<100:
            nrFoto=str(nrFoto)
            nomeFoto="Foto000"+nrFoto+"."+"jpg"
            print("il numero della foto è da 0 a 99")
            nrFoto=int(nrFoto)
        if nrFoto>99 and nrFoto<1000:
            nrFoto=str(nrFoto)
            nomeFoto="Foto00"+nrFoto+"."+"jpg"
            nrFoto=int(nrFoto)
            print("il numero della foto è da 100 a 999")
        if nrFoto>999 and nrFoto<10000:
            print("il numero della foto è da 999 a 9999")
            nrFoto=str(nrFoto)
            nomeFoto="Foto0"+nrFoto+"."+"jpg"
            nrFoto=int(nrFoto)
        if nrFoto>9999 and nrFoto<100000:
            print("il numero della foto è da 9999 a 99999")
            nrFoto=str(nrFoto)
            nomeFoto="Foto"+nrFoto+"."+"jpg"
            nrFoto=str(nrFoto)   

        indirNomeFoto=""+nomeFoto
        print("il nome completo della foto è ",indirNomeFoto)
        vecchioNome = "/home/pi/plant+out/PlantToWeb/ultima.jpg"
        nuovoNome = "/home/pi/plant+out/PlantToWeb/"+nomeFoto
        path1="/home/pi/plant+out/PlantToWeb/"+nomeFoto
        path2="/media/pi/45D5-C1E61/FotoDiPiante"
        
        os.rename(vecchioNome,nuovoNome)
        shutil.move(path1,path2)
        print("ora ultima.jpg si trova qui ",nuovoNome)
        #a questo punto la cartella /home/pi/plant+out/PlantToWeb si ritrova vuota
        #verificare se va bene o se è meglio un shutil.copy al posto di shutil.move

        

def scriviDatiSuGoogleFogli():
    import openpyxl
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx')
    ws=wb["Sheet1"]
    #ws1=wb["ultime48ore"]
    #ws2=wb["ultimaSettimana"]
    #ws3=wb["ultime2settimane"]
    #ws4=wb["ultimoMese"]
    #ws5=wb["datiGiornalieri"]
    #ws8=wb["datiorari"]
#inserisci i valori copiati da store
#ws=wb[Sheet1] non funziona non so perchè
#cell=ws['A328'].value
#provo ad estrarre un valore dal foglio
#print cell
#provo a contare le righe
    r=ws.max_row
    print (r)
#a questo punto estraggo la dataora dell'ultimo rilievo:
    ora=ws.cell(row=r,column=1).value
    print (ora)
#ora provo ad estrarre la temperatura del suolo
    tsuolo1=ws.cell(row=r,column=6).value
    tsuolo2=ws.cell(row=r,column=8).value
#print tsuolo
#ora provo ad estrarre la luce alla pianta
    #lux=ws.cell(row=r,column=4).value#controllare se questo lo posso eliminare
#print lux
#ora provo ad estrarre la temperatura dell'aria
    taria=ws.cell(row=r,column=2).value
#print taria
#ora provo ad estrarre l'umidità relativa
    RH=ws.cell(row=r,column=3).value
#print RH
#ora provo ad estrarre la bagnatura del suolo
    Bagnatura1=ws.cell(row=r,column=5).value
    Bagnatura2=ws.cell(row=r,column=7).value
#print res
#ora provo ad estrarre la luce ambiente
    luce=ws.cell(row=r,column=4).value
#print l_amb
#continua ad estrarre tutti i valori con i nomi che devono essere inseriti in google fogli
#inserisco i valori da store nel foglio prove  di fogli Google
    wks.append_row([ora,taria, RH, luce, Bagnatura1,tsuolo1,Bagnatura2,tsuolo2])
    print("ho scritto i dati in GoogleSheet RpiPlantOut1Logger, foglio sheet1")
    print("al seguente indirizzo: https://docs.google.com/spreadsheets/d/16mNIVRM7lZpjAZivV8HqeAFI-WcwHP0H5DwMsHth1Yw/edit#gid=0")
    #time.sleep(2)
    del openpyxl
    
#ho cambiato le funzioni in modo che su store vengano aggiornati altri fogli: 48 0re (con la media oraria), dati giornalieri, 1 settimana, 2 settimane, 1 mese
#poi copio tutto su googleSheet
#per fare questo ho cambiato gli step 7 - 8 - 9 che ho messo nella cartella scuciECuci
#che devo sostituire nel file tutto 
    
#qui inserire le funzioni della scrittura di altri fogli di store
def scrivi48oreStore():
    import openpyxl        
    #wb=openpyxl.load_workbook('/home/pasquale/Scrivania/StoreProve.xlsx',read_only=False)
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
    ws=wb["Sheet1"]
    ws1=wb["ultime48ore"]
    ws8=wb["datiOrari"]
    import datetime
    from datetime import datetime
    import time
    #la funzione aggiorna anche il foglio dati orari su store(dalla versione 24)
    ultimaRigaSheet1=ws.max_row
    print("l'ultima riga del foglio 1 di store e' ",ultimaRigaSheet1)
    ultima48ore=ws1.max_row
    print("l'ultima riga del foglio 48 ore e' ",ultima48ore)
    time.sleep(15)
    if ultimaRigaSheet1<5:
        print("ci sono meno di 4 dati scritti nel foglio1 di store non fare nessun calcolo, termina questa funzione e proseguo altre")
        del openpyxl
    elif ultimaRigaSheet1>=5 and ultima48ore<49:
        print("attiva la parte di script che fa la media oraria e aggiunge i dati della media oraria al foglio 48 ore")
        #ho incollato qui il contenuto della funzione mediaDatiSuStore
        ultimaRigaSheet1=ws.max_row
        print("l'ultima riga scritta nello sheet1 nel file store è ", ultimaRigaSheet1)
        nc=9#indichiamo il numero di colonne per il quale deve essere ripetuta la media
        mediaValori=[]#è una lista che conterrà tutte le medie dei valori
        for s in range (2,nc):#s=il numero della colonna
            l=[]#creo una lista vuota
            n=(ws.max_row-3)
            print (n, " è la riga dove iniziare la media dei valori")
            print ("il contenuto della prima cella nella prima colonna dove iniziare la media dei valori è ", ws.cell(row=n, column=2).value)         
            for d in range (4):#per 4 volte
                print("la riga di lettura attuale è ",n)
                v=ws.cell(row=n,column=s).value #estrae il valore della prima riga della colonna s e lo chiama v
                #v=cell.value
                if v!=None:
                    v=float(v)#lo converto in numero altrimenti non fa i calcoli
                else:
                    v=0
                    #print("il valore della cella estratta alla riga ", n," è ", v)
                l[d:d]=[v] #dovrebbe inserirlo nell lista, si chiama l che contiene i 4 valori rilevati
                n=n+1
                d=d+1#sposta la posizione nella lista di 1               
            print(l)
            a=sum(l)/len(l)#chiamo a la media dei valori nella lista l
            a=round(a,2) #arrotondo la media a 2 cifre
            print("la media dei valori di ", ws.cell(row=1, column=s).value," degli ultimi 4 rilievi è ",a)
            print("ora vado a fare i calcoli della colonna ",s)
            #time.sleep(2)
            mediaValori.append(a)        
        print("La media oraria dei valori rilevati è ",mediaValori)
        dataOra=ws.cell(row=(ultimaRigaSheet1-1),column=1).value
        #print (dataOra)
        #print("data è ",type(dataOra))
        dataOra=datetime.strptime(dataOra, "%d/%m/%Y %H:%M:%S")
        #print (dataOra)
        dataOra=dataOra.strftime('%d/%m/%Y %H')
        print ("data ora con strftime è ",dataOra)
        #print (type(dataOra))
        #media valori contiene tutte le medie orari dei valori ma non data e ora
        #definisco tutti i valori medi estraendoli dalla lista
        TAmb=mediaValori[0]
        UR=mediaValori[1]
        Luce=mediaValori[2]
        BS10=mediaValori[3]
        TS10=mediaValori[4]
        BS30=mediaValori[5]
        TS30=mediaValori[6]
        #Aggiungo la riga della data e ora e i valori medi ai fogli(48 ore e dati orari)
        ws1.append([dataOra, TAmb, UR, Luce, BS10, TS10, BS30, TS30])
        ws8.append([dataOra, TAmb, UR, Luce, BS10, TS10, BS30, TS30])#aggiorno dati orari
        print("ho aggiornato il foglio 48 ore e dati giornalieri")
        #wb.save("/home/pasquale/Scrivania/StoreProve.xlsx")
        wb.save("/home/pi/plant+out/store.xlsx")
        print("ho salvato store")
        del openpyxl
    elif ultimaRigaSheet1>=5 and ultima48ore==49:
        print("attiva la parte di script che fa la media oraria e in 48ore sposta i dati un rigo più in alto ed inserisce la media appena fatta al rigo 49")
        ws1.move_range("A3:H49", rows=-1, cols=0)
        print("ho spostato una riga più in alto il blocco dati in 48ore")
        print("attiva la parte di script che fa la media oraria e aggiunge i dati della media oraria al foglio 48 ore")
        ultimaRigaSheet1=ws.max_row
        print("l'ultima riga scritta nello sheet1 nel file store è ", ultimaRigaSheet1)
        nc=9#indichiamo il numero di colonne per il quale deve essere ripetuta la media
        mediaValori=[]#è una lista che conterrà tutte le medie dei valori
        for s in range (2,nc):#s=il numero della colonna
            l=[]#creo una lista vuota
            n=(ws.max_row-3)
            print (n, " è la riga dove iniziare la media dei valori")
            print ("il contenuto della prima cella nella prima colonna dove iniziare la media dei valori è ", ws.cell(row=n, column=2).value)         
            for d in range (4):#per 4 volte
                print("la riga di lettura attuale è ",n)
                v=ws.cell(row=n,column=s).value #estrae il valore della prima riga della colonna s e lo chiama v
                #v=cell.value
                if v!=None:
                    v=float(v)#lo converto in numero altrimenti non fa i calcoli
                else:
                    v=0
                    #print("il valore della cella estratta alla riga ", n," è ", v)
                l[d:d]=[v] #dovrebbe inserirlo nell lista, si chiama l che contiene i 4 valori rilevati
                n=n+1
                d=d+1#sposta la posizione nella lista di 1               
            print(l)
            a=sum(l)/len(l)#chiamo a la media dei valori nella lista l
            a=round(a,2) #arrotondo la media a 2 cifre
            print("la media dei valori di ", ws.cell(row=1, column=s).value," degli ultimi 4 rilievi è ",a)
            print("ora vado a fare i calcoli della colonna ",s)
            #time.sleep(2)
            mediaValori.append(a)        
        print("La media oraria dei valori rilevati è ",mediaValori)
        dataOra=ws.cell(row=(ultimaRigaSheet1-1),column=1).value
        #print (dataOra)
        #print("data è ",type(dataOra))
        dataOra=datetime.strptime(dataOra, "%d/%m/%Y %H:%M:%S")
        #print (dataOra)
        dataOra=dataOra.strftime('%d/%m/%Y %H')
        print ("data ora con strftime è ",dataOra)
        #print (type(dataOra))
        #media valori contiene tutte le medie orari dei valori ma non data e ora
        #definisco tutti i valori medi estraendoli dalla lista
        TAmb=mediaValori[0]
        UR=mediaValori[1]
        Luce=mediaValori[2]
        BS10=mediaValori[3]
        TS10=mediaValori[4]
        BS30=mediaValori[5]
        TS30=mediaValori[6]

        ws1['A49'] = dataOra
        ws1['B49'] = TAmb
        ws1['C49'] = UR
        ws1['D49'] = Luce
        ws1['E49'] = BS10
        ws1['F49'] = TS10
        ws1['G49'] = BS30
        ws1['H49'] = TS30
        print("ho aggiornato il foglio 48 ore")
        ws8.append([dataOra, TAmb, UR, Luce, BS10, TS10, BS30, TS30])#aggiorno dati orari
        # cambiare quanto segue perchÉscrive su rigo 49 e non ultimo tipo ws8.append([dataOra, TAmb, UR, Luce, BS10, TS10, BS30, TS30])#aggiorno dati orari
        """
        ws8['A49'] = dataOra
        ws8['B49'] = TAmb
        ws8['C49'] = UR
        ws8['D49'] = Luce
        ws8['E49'] = BS10
        ws8['F49'] = TS10
        ws8['G49'] = BS30
        ws8['H49'] = TS30

        print("ho aggiornato  dati orari")
        """
        print("dovrei avere aggiornato i dati orari: ws8")
        #wb.save("/home/pasquale/Scrivania/StoreProve.xlsx")
        wb.save("/home/pi/plant+out/store.xlsx")
        print("ho salvato store")
        del openpyxl

def mediaDatiGiornalieri():
    import openpyxl        
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
    ws1=wb["ultime48ore"]
    ws5=wb["datiGiornalieri"]
    ultima48ore=ws1.max_row
    if ultima48ore<25:
        print("Ci sono meno di 24 dati orari perché l'ultima riga è ", ultima48ore," in foglio 48 ore. Non verrà eseguita la media giornaliera")
    else:
        dataOraG=ws1.cell(row=ultima48ore, column=1).value#questa va bene, in effetti no, nom mi piace devo togliere l'ora
        data=dataOraG[0:10]
        print("la data da riportare nel foglio datiGiornalieri è ", data)
        print("il formato della data è ",type(data))#èuna stringa la devo trasformare in data
        #data=datetime.strptime(data, "%d/%m/%Y") #e ma se la trasformo in data mi rimette l'ora: la lascio così
        #print(data)
        #print("il formato della data è ",type(data))
        #Per gli altri dati faccio un ciclo for inizio da row 26
        #mutuo dall'altro calcolo della media lì iterava 4 volte sulle righe qui 24 (ore del giorno)
        nc=9#indichiamo il numero di colonne per il quale deve essere ripetuta la media
        maxValori=[]
        mediaValori=[]
        minValori=[]
        for s in range (2,nc):#s=il numero della colonna
            print("la colonna su cui faccio i calcoli ora è ",s)
            l=[]#creo una lista vuota per i valori medi che chiam l
            n=ultima48ore-23
            print (n, " è la riga dove iniziare la media dei valori")            
            for d in range (24):#per 24 volte
                print("la riga di lettura attuale è ",n)
                v=ws1.cell(row=n,column=s).value #estrae il valore della prima riga della colonna s e lo chiama v
                #v=cell.value
                if v!=None:
                    v=float(v)#lo converto in numero altrimenti non fa i calcoli
                else:
                    v=0
                    #print("il valore della cella estratta alla riga ", n," è ", v)
                l[d:d]=[v] #dovrebbe inserirlo nell lista
                n=n+1
                d=d+1#sposta la posizione nella lista di 1    
            #time.sleep(1)   
            print("i valori della colonna ", s," delle ultime 24 ore sono: ",l)
            a=sum(l)/len(l)#chiamo a la media dei valori nella lista
            a=round(a,2) #arrotondo la media a 2 cifre
            print("la media dei valori di ", ws1.cell(row=1, column=s).value," degli ultimi 24 rilievi è ",a)
            b=max(l)
            print("il massimo valore di ", ws1.cell(row=1, column=s).value," degli ultimi 24 rilievi è ",b)
            c=min(l)
            print("il minimo valore di ", ws1.cell(row=1, column=s).value," degli ultimi 24 rilievi è ",c)
            #time.sleep(1)
            #time.sleep(2)
            mediaValori.append(a)
            maxValori.append(b)
            minValori.append(c)
        print("La media oraria dei valori rilevati è ",mediaValori)
        print("Il massimo dei dati orari dei valori rilevati nelle ultime 24 ore è ",maxValori)
        print("Il minimo dei dati orari dei valori rilevati nelle ultime 24 ore è ",minValori)      
        print("La media oraria dei valori rilevati è ",mediaValori)
        TAmbG=mediaValori[0]
        URG=mediaValori[1]
        LuceG=mediaValori[2]
        BS10G=mediaValori[3]
        TS10G=mediaValori[4]
        BS30G=mediaValori[5]
        TS30G=mediaValori[6]
        
        TAmbMG=maxValori[0]
        URMG=maxValori[1]
        LuceMG=maxValori[2]
        BS10MG=maxValori[3]
        TS10MG=maxValori[4]
        BS30MG=maxValori[5]
        TS30MG=maxValori[6]
        
        TAmbmG=minValori[0]
        URmG=minValori[1]
        LucemG=minValori[2]
        BS10mG=minValori[3]
        TS10mG=minValori[4]
        BS30mG=minValori[5]
        TS30mG=minValori[6]        
        ws5.append([data,TAmbG,URG,LuceG,BS10G,TS10G,BS30G,TS30G,TAmbMG,TAmbmG,URMG,URmG,LuceMG,LucemG,BS10MG,BS10mG,TS10MG,TS10mG,BS30MG,BS30mG,TS30MG,TS30mG])
        print("l'ultima riga scritta nei dati giornalieri è ",ws5.max_row)
        #time.sleep(1)
        wb.save("/home/pi/plant+out/store.xlsx")
        print("l'ultima riga scritta nei dati giornalieri dopo il salvataggio è ",ws5.max_row)
        #time.sleep(1)
        wb.close()
        print ("ho salvato la media di tutti i dati giornalieri su store...")

        wb.close
        del openpyxl
    


def aggiornaDatiGiornalieriSuGSheet():       
    import openpyxl        
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
    ws5=wb["datiGiornalieri"]
    print("l'ultima riga scritta nei dati giornalieri è ",ws5.max_row)
    ultimaRigaDG=ws5.max_row
    
    data=ws5['A'+str(ultimaRigaDG)].value
    TAmbG=ws5['B'+str(ultimaRigaDG)].value
    URG=ws5['C'+str(ultimaRigaDG)].value
    LuceG=ws5['D'+str(ultimaRigaDG)].value
    BS10G=ws5['E'+str(ultimaRigaDG)].value
    TS10G=ws5['F'+str(ultimaRigaDG)].value
    BS30G=ws5['G'+str(ultimaRigaDG)].value
    TS30G=ws5['H'+str(ultimaRigaDG)].value
    MTArG=ws5['I'+str(ultimaRigaDG)].value
    mTArG=ws5['J'+str(ultimaRigaDG)].value
    MURG=ws5['K'+str(ultimaRigaDG)].value
    mURG=ws5['L'+str(ultimaRigaDG)].value
    MLuceG=ws5['M'+str(ultimaRigaDG)].value
    mLuceG=ws5['N'+str(ultimaRigaDG)].value
    MBS10G=ws5['O'+str(ultimaRigaDG)].value
    mBS10G=ws5['P'+str(ultimaRigaDG)].value
    MTS10G=ws5['Q'+str(ultimaRigaDG)].value
    mTS10G=ws5['R'+str(ultimaRigaDG)].value
    MBS30G=ws5['S'+str(ultimaRigaDG)].value
    mBS30G=ws5['T'+str(ultimaRigaDG)].value
    MTS30G=ws5['U'+str(ultimaRigaDG)].value
    mTS30G=ws5['V'+str(ultimaRigaDG)].value
    
    
    print("Gli ultimi dati giornalieri sono ",data,TAmbG,URG,LuceG,BS10G,TS10G,BS30G,TS30G,
    MTArG,mTArG,MURG,mURG,MLuceG,mLuceG,MBS10G,mBS10G,MTS10G,mTS10G,MBS30G,mBS30G,MTS30G,mTS30G,
    "e li sto scrivendo in Gsheet")
    wks4.append_row([data,TAmbG,URG,LuceG,BS10G,TS10G,BS30G,TS30G,
    MTArG,mTArG,MURG,mURG,MLuceG,mLuceG,MBS10G,mBS10G,MTS10G,mTS10G,MBS30G,mBS30G,MTS30G,mTS30G,])#aggiorno il foglio dati giornalieri su gsheet
    print("... e ora ho scritto tutte le medie dei dati giornalieri anche su gSheet")
    #continua con l'aggiornamento di altri fogli

 
def scrivoDatiGiornalieriInAltriFogli(foglio, nMaxRighe):
    """
    import openpyxl
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
    ws5=wb["datiGiornalieri"]
    ws2=wb["ultimaSettimana"]
    ws3=wb["ultime2settimane"]
    ws4=wb["ultimoMese"]
    """
    #anche queste funzioni sono da lanciare a mezzanotte
    #estraggo i dati dal foglio dati giornalieri
    urdg=ws5.max_row #ultima riga dati giornalieri
    print("l'ultima riga del foglio dati giornalieri è ", urdg)
    #time.sleep(1)
    dataOraG=ws5.cell(row=urdg,column=1).value 
    TAmbG=ws5.cell(row=urdg,column=2).value
    URG=ws5.cell(row=urdg,column=3).value
    LuceG=ws5.cell(row=urdg,column=4).value
    BS10G=ws5.cell(row=urdg,column=5).value
    TS10G=ws5.cell(row=urdg,column=6).value
    BS30G=ws5.cell(row=urdg,column=7).value
    TS30G=ws5.cell(row=urdg,column=8).value
    MTArG=ws5.cell(row=urdg,column=9).value
    mTArG=ws5.cell(row=urdg,column=10).value
    MURG=ws5.cell(row=urdg,column=11).value
    mURG=ws5.cell(row=urdg,column=12).value
    MLuceG=ws5.cell(row=urdg,column=13).value
    mLuceG=ws5.cell(row=urdg,column=14).value
    MBS10G=ws5.cell(row=urdg,column=15).value
    mBS10G=ws5.cell(row=urdg,column=16).value
    MTS10G=ws5.cell(row=urdg,column=17).value
    mTS10G=ws5.cell(row=urdg,column=18).value
    MBS30G=ws5.cell(row=urdg,column=19).value
    mBS30G=ws5.cell(row=urdg,column=20).value
    MTS30G=ws5.cell(row=urdg,column=21).value
    mTS30G=ws5.cell(row=urdg,column=22).value
    
    if  urdg<2:
        print("Non ci sono ancora valori in 'datiGiornalieri' non verranno copiati i dati giornalieri in altri fogli")
        #del openpyxl
        
    elif foglio.max_row<nMaxRighe:
        print ("il numero di righe in ",foglio, "è ", foglio.max_row," e pertanto minore di ", nMaxRighe, ": aggiungo semplicemente i dati giornalieri")
        foglio.append([dataOraG, TAmbG, URG, LuceG, BS10G, TS10G, BS30G, TS30G,
        MTArG,mTArG,MURG,mURG,MLuceG,mLuceG,MBS10G,mBS10G,MTS10G,mTS10G,
        MBS30G,mBS30G, MTS30G,mTS30G])
        print ("i dati giornalieri sono: ", dataOraG, TAmbG, URG, LuceG, BS10G, TS10G, BS30G, TS30G,
        MTArG,mTArG,MURG,mURG,MLuceG,mLuceG,MBS10G,mBS10G,MTS10G,mTS10G,
        MBS30G,mBS30G, MTS30G,mTS30G," e sono stati inseriti alla riga ",foglio.max_row+1)
        #time.sleep(1)
        wb.save("/home/pi/plant+out/store.xlsx")
        #wb.close()
        print("il file store con ", foglio, "aggiornato nella directory plant+out")
        print("l'aggiornamento dei dati giornalieri in ",foglio, "è stato completato")
        #del openpyxl
        
    elif foglio.max_row>nMaxRighe-1:
        print ("il numero di righe in ",foglio, "è maggiore di ", nMaxRighe-1, "pertanto sposto i dati di una riga in alto ed inserisco i dati alla riga ",nMaxRighe )
        #time.sleep(1)
        foglio.move_range("A3:V"+str(nMaxRighe),rows=-1)
        foglio["A"+str(nMaxRighe)]=dataOraG
        foglio["B"+str(nMaxRighe)]=TAmbG
        foglio["C"+str(nMaxRighe)]=URG
        foglio["D"+str(nMaxRighe)]=LuceG
        foglio["E"+str(nMaxRighe)]=BS10G
        foglio["F"+str(nMaxRighe)]=TS10G
        foglio["G"+str(nMaxRighe)]=BS30G
        foglio["H"+str(nMaxRighe)]=TS30G
        foglio["I"+str(nMaxRighe)]=MTArG
        foglio["J"+str(nMaxRighe)]=mTArG
        foglio["K"+str(nMaxRighe)]=MURG
        foglio["L"+str(nMaxRighe)]=mURG
        foglio["M"+str(nMaxRighe)]=MLuceG
        foglio["N"+str(nMaxRighe)]=mLuceG
        foglio["O"+str(nMaxRighe)]=MBS10G
        foglio["P"+str(nMaxRighe)]=mBS10G
        foglio["Q"+str(nMaxRighe)]=MTS10G
        foglio["R"+str(nMaxRighe)]=mTS10G
        foglio["S"+str(nMaxRighe)]=MBS30G
        foglio["T"+str(nMaxRighe)]=mBS30G
        foglio["U"+str(nMaxRighe)]=MTS30G
        foglio["V"+str(nMaxRighe)]=mTS30G
        print("i dati giornalieri sono ", dataOraG, TAmbG, URG, LuceG, BS10G, TS10G, BS30G, TS30G,
        MTArG,mTArG,MURG,mURG,MLuceG,mLuceG,MBS10G,mBS10G,MTS10G,mTS10G, MBS30G,mBS30G, MTS30G,mTS30G)
        print("ho inserito i dati alla riga ",nMaxRighe)
        #time.sleep(1)
        wb.save("/home/pi/plant+out/store.xlsx")
        #wb.close()
        #del openpyxl
        print("il file store con ", foglio, "aggiornato nella directory plant+out")
        print("l'aggiornamento dei dati giornalieri in ",foglio, "è stato completato")
        








def daStoreAGsheet(wsstore,wsgsheet,colonnastore):#devo sostituire i valori alle variabili della funzione    
    rigastore=wsstore.max_row
    list=[]#creo una lista principale vuota il cui indice sarà h. E' una lista che contiene tante liste quante sono le righe
    h=0 #Per metterlo in ordine giusto
    for n in range(2,rigastore+1):#faccio un ciclo che ripete le azioni per ogni riga   
        lista=[]#creo una lista secondaria vuota il cui indice sarà i
        for m in range (2,colonnastore+1): #il ciclo per ogni colonna    
            i=colonnastore-1 #n. colonne - 1 - Per metterlo in ordine giusto nella lista
            #parto dall'indice più alto (non ho capito bene perchè)
            #print(m)
            dato=wsstore.cell(row=n,column=m).value
            if dato==None:
                dato=0
                print("il valore era None e l'ho traformato in ",dato)
            else:
                dato=dato
                print("il valore non era None ma ",dato," e l'ho lasciato tale")
            lista[i:i]=[dato]
            i=i-1
        list[h:h]=[lista]
        h=h+1
    
    print(list)#così ho una lista con tutti i valori dentro

    dati=list#non ho capito bene il significato di questa variabile
    #print (dati)

    #In effetti non ho capito bene come funziona
    #pertanto non so cambiarlo, ad esempio non so come fare affinché ci sia una lista che contiene
    #tante liste quante sono le colonne
    #devo seguire lalogica della lista madre che contiene tante liste quante sono le righe
    print(len(list))
    for i in range(len(list)):#sono gli indici per ogni riga
        for j in range(len(dati[i])):#sono gli indici ogni colonna
            dati[i][j] = float(dati[i][j])#trasformo i dati da str a float di ogni cella del range di righe e colonne
    print(dati)#per capire che effettivamente siano float
    #funziona, siccome non sono riuscito a cambiare la nidificazione (colonne dentro righe) nella
    #scrittura dei dati gSheet allora ho cambiato l'estrazione dei valori da store
    wsgsheet.update('B2', dati)#i dati nelle liste interne le inserisce nelle righe, le liste esterne le inserisce nelle colonne

    #ora cerco di copiare le date in gsheet
    
    date1=[]
    for d in range(2,rigastore+1):
        g=rigastore-2
        data=wsstore.cell(row=d,column=1).value
        date1[g:g]=[data]
        i=g-1
    print(date1)
    date=[]
    date[1:1]=[date1]
    print(date)
    dateG=wsgsheet.get('A2:A'+str(rigastore+1))
    print(dateG)
    #in gSheet per scrivere i dati in una riga i valori devono essere contenuti in una lista
    #per scrivere una colonna la lista deve contenere altre liste con i valori
    #pertanto fare una lista date con indici di Posizione variabil e scrivere in questa la lista date1 che contengono le date con indice di posizione sempre 1

    date=[]#creo una lista esterna vuota date
    g=1#indice di posizione di date che è variabile
    for d in range (2,rigastore+1):
        date1=[]#queste sono le liste interne che contengono un solo dato(indice di posizione sempre 1)
        data=wsstore.cell(row=d,column=1).value#estraggo il valore dalle celle
        date1[1:1]=[data]# lo metto nella lista interna
        #print(date1)
        date[g:g]=[date1]#metto la lista interna in quella esterna in posizione g
        g=g+1#vario la posizione nella lista esterna
        #print(date)
    print(date)
    wsgsheet.update('A2', date)
    print("ho copiato i dati da store", wsstore," in ", wsgsheet, " di gSheet")
    
def creaVideo():
    import os
    import time
    import datetime
    import openpyxl
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx')
    ws6=wb["videoInLocale"]#dst in colonna A; ora in colonna B
    
    now=datetime.datetime.now()# chiedo che ora è
    ora=now
    print ("ora è proprio ...")
    print (ora)
    #time.sleep(1)
    #scrivo la data e l'ora che poi mi serve per rinominare la nuova
    #cartella dove viene dopositato il video con le foto
    ora= ora.strftime('%d-%m-%Y_%He%M')
    print ("ora formattato nella maniera giusta e' "+ora)

    #time.sleep(1)
    prefix="video"
    newDir=prefix+ora
    print ("la cartella dove verranno spostate le foto e il video e' "+newDir)

    #ho dovuto costruire questa nuova variabile per inserirla in rename
    #dst = la destinazione

    dst='/media/pi/45D5-C1E61/'+newDir
    #time.sleep(1)
    print ("l'indirizzo assoluto della nuova cartella è "+dst)
    print ("devo depositare i dati dst e ora nel file store.xsls in quanto verrà poi riutilizzato nella funzione \n mandoVideoSuWeb che potrebbe essere lanciata in un secondo momento")
    ws6.append([dst, ora])
    print("dovrei avere scritto ",dst, " e ", ora , " in foglio videoInLocale di store.xsls")
    wb.save('/home/pi/plant+out/store.xlsx')
    wb.close()
    del openpyxl
    
    
    #ora devo rinominare fotoDiPiante con newDir 
    os.rename('/media/pi/45D5-C1E61/FotoDiPiante', dst)
    #time.sleep(1)
    print ("a questo punto ho rinominato FotoDiPiante in "+newDir)


    #una volta stabilito che "ora" e' quel dato momento la varibile rimane tale
    #anche al variare del tempo questo mi consente poi di recuperare il file
    #i file nella directory
    #time.sleep(1)
    #ora devo fare una nuova cartella FotoDiPiante che risultera' vuota
    #e che immagazzinerà nuove foto dopo il video.
    print ("creo una nuova cartella FotoDiPiante vuota dove verranno messe le foto dopo la creazione del video")
    os.mkdir('/media/pi/45D5-C1E61/FotoDiPiante') #si possono specificare anche i permessi ma al momento
                                                #non ho capito come si fa, ma li mette di default 777
    #dst è l'indirizzo assoluto dove ci sono le foto per montare il video e dove verrà depositato il video
    cmd="ffmpeg -f image2 -r 10 -i "+dst+"/Foto%05d.jpg -vf scale=iw/1.8:-1 -c:v libx264 -profile:v high -level 5 -crf 23 -maxrate 10M -bufsize 16M -pix_fmt yuv420p -x264opts bframes=3:cabac=1 -movflags faststart "+dst+"/movie"+ora+".mp4"
    time.sleep(1)
    print ("il comando shell per la creazione del video è il seguente: "+cmd)
    time.sleep(1)
    print ("inizio della creazione del video.... c'impiegherà qualche minuto")
    print (os.system(cmd))
    #wow funziona! Python is cool!!!

    print (".....dovrebbe essere finita la creazione del video")

    time.sleep(1)
    print ("a questo punto la funzione termina perchÉ c'è un'altra funzione che manda il video su web")
    print("pertanto fino a questo punto ho creato il video  che si  chiama movie"+ora+".mp4 ","ed e è depositato in questa cartella ", dst)

    

#la funzione che segue va lanciata solo dopo la verifica della connessione
def mandaVideoSuWeb():
    print ("ora provo a lanciare il video su web")
    #continua con l'invio del video
    #per prima cosa devo capire come aspettare che la ostruzione del video finisca.
    #poi lo devo inviare su web. Ma non vorrei farlo lanciando il file bash.
    #in quanto mi piacerebbe lanciare il video su web in forma personale con nome video+
    #data e ora.avi Pertanto sarebbe importante che il cmando bash prendesse le variabili
    #Potrei provare a lanciare in python con os.system un comando alla volta del fle bash
    import time
    import ftplib
   
    ultimaRiga=ws6.max_row
    print(ultimaRiga)
    dst=ws6['A'+str(ultimaRiga)].value
    print(dst)
    ora=ws6['B'+str(ultimaRiga)].value
    print (ora)

    session=ftplib.FTP('rpiplant.altervista.org', 'rpiplant','aPQq7jNJpWzS')
    nomefile=dst+"/movie"+ora+".mp4"
    nomefile1="movie"+ora+".mp4"
    print(nomefile)
    print(nomefile1)
    session.storbinary("STOR " + nomefile1, open (nomefile, "rb"), 1024)
    session.quit()

    time.sleep(1)
    print ("il filmato dovrebbe essere stato inviato su rpiplant.altervista. si dovrebbe trovare lì con questo nome  "+nomefile1, " a questo indirizzo www.rpiplant.altervista.org/"+nomefile1)
    body="www.rpiplant.altervista.org/"+nomefile1
    time.sleep(1)
    print("scrivo l'indirizzo del video appena prodotto su un file txt che si chiama videoInWeb .... non so perchè")
    videoInWeb = open('/home/pi/plant+out/videoInWeb.txt','w')
    videoInWeb.write(body)
    videoInWeb.close()
    wks9.append_row([body])
    print("dovrei aver scritto l'indirizzo del video anche in gSheet")
    ws7.append([body])
    wb.save('/home/pi/plant+out/store.xlsx')
    wb.close()
    
    print("ho scritto l'indirizzo del video anche in locale su store")
   
    
def sendAlarmBagnaturaSuolo():   
    import openpyxl
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx')
    ws=wb["Sheet1"]
    #provo ad estrarre il valore dell'ultima bagnatura dal foglio
    #provo a contare le righe
    r=ws.max_row
    print( r)
    #a questo punto estraggo la dataora dell'ultimo rilievo:
    if r<24:
        print(" ci sono pochi dati di rilievo dell'umidità del suolo per inviare l'allarme")
    else:
        print("i rilievi  dell'umidità del suolo sono sufficienti per la verifica dell'invio allarme")
        bs1=ws.cell(row=r,column=5).value
        print (bs1)
        #faccio un ciclo for da 1 a 48
        BS1=[]#creo una lista vuota
        for i in range(24): #media delle ultime  6 ore
            p=0#posizione nella lista
            bs1=ws.cell(row=r,column=5).value
            #print bs
            r=r-1#scorro una riga di lettura indietro del file store
            #Ok ora però li devo mettere in una lista
            BS1[p:p]=[bs1]#inserisco il valore nella lista
            p=p+1#scorro di una posizione
            print(bs1)

        #print BS
        a=sum(BS1)
        b=len(BS1)
        bsm1=a/b
        bsm1=round(bsm1,0)
        #print a
        #print b
        print ("la media della bagnatura del suolo del sensore superficiale delle ultime 6 ore è ",bsm1)
        
        BS2=[]#creo una lista vuota
        for i in range(24): #media delle ultime  6 ore
            p=0#posizione nella lista
            bs2=ws.cell(row=r,column=7).value
            #print bs
            r=r-1#scorro una riga di lettura indietro del file store
            #Ok ora però li devo mettere in una lista
            BS2[p:p]=[bs2]#inserisco il valore nella lista
            p=p+1#scorro di una posizione
            print(bs2)

        #print BS
        a=sum(BS2)
        b=len(BS2)
        bsm2=a/b
        bsm2=round(bsm2,0)
        #print a
        #print b
        print ("la media della bagnatura del suolo del sensore profondo delle ultime 6 ore è ",bsm2)

        #Ora devo confrontare il valore medio della bagnatura nelle 6 ore con  paramentri
        #estraggo i paramentri in rpiPlantOut1Logger 
        

        LimEccIdr1=wks13.acell('B9').value#limite al di sopra del quale il sensore superficiale rileva un eccesso idrico
        LimEccIdr2=wks13.acell('B10').value#limite al di sopra del quale il sensore profondo rileva un eccesso idrico

        LimDefIdr1=wks13.acell('C9').value#limite al di sotto del quale il sensore superficiale rileva un deficit idrico
        LimDefIdr2=wks13.acell('C10').value#limite al di sotto del quale il sensore profondo rileva un deficit idrico

        LimStrIdr1=wks13.acell('D9').value#limite al di sotto del quale il sensore superficiale rileva uno stress idrico
        LimStrIdr2=wks13.acell('D10').value#limite al di sotto del quale il sensore profondo rileva uno stress idrico

        #trasformo tutto in float
        LimEccIdr1=float(LimEccIdr1)
        LimEccIdr2=float(LimEccIdr2)

        LimDefIdr1=float(LimDefIdr1)
        LimDefIdr2=float(LimDefIdr2)

        LimStrIdr1=float(LimStrIdr1)
        LimStrIdr2=float(LimStrIdr2)

        print("il limite al di sopra del quale il sensore superficiale rileva un eccesso idrico è ",LimEccIdr1)
        print("il limite al di sopra del quale il sensore profondo rileva un eccesso idrico è ", LimEccIdr2)
        print("il limite al di sotto del quale il sensore superficiale rileva un deficit idrico è ",LimDefIdr1)
        print("il limite al di sotto del quale il sensore profondo rileva un deficit idrico è ", LimDefIdr2)
        print("il limite al di sotto del quale il sensore superficiale rileva uno stress idrico è ", LimStrIdr1)
        print("il limite al di sotto del quale il sensore profondo rileva uno stress idrico è ", LimStrIdr2)

        #definizione di oggetto

        if bsm1<LimStrIdr1 or bsm2<LimStrIdr2:
            oggetto='Stress idrico - urgente irrigare: valore del sensore superficiale= '+str(bsm1)+' valore del sensore profondo= '+str(bsm2)
            print (oggetto)
        elif (bsm1>LimStrIdr1 and bsm1<LimDefIdr1) or (bsm2>LimStrIdr2 and bsm2<LimDefIdr2):
            oggetto='deficit idrico - necessario irrigare: valore del sensore superficiale= '+str(bsm1)+' valore del sensore profondo= '+str(bsm2)
            print (oggetto)
        elif bsm1>LimEccIdr1 or bsm2>LimEccIdr2:
            oggetto='acqua in eccesso: valore del sensore superficiale= '+str(bsm1)+' valore del sensore profondo= '+str(bsm2)
            print (oggetto)
        else:
            print ("stato idrico ok: nessun invio")

        eMail1=wks13.acell('B4').value
        eMail2=wks13.acell('B5').value
        eMail3=wks13.acell('B6').value

        print("la prima mail a cui spedire gli allarmi relativi al contenuto idrco del suolo è ", eMail1)
        print("la seconda mail a cui spedire gli allarmi relativi al contenuto idrco del suolo è ", eMail2)
        print("la terza mail a cui spedire gli allarmi relativi al contenuto idrco del suolo è ", eMail3)

        eMail1=str(eMail1)
        eMail2=str(eMail2)
        eMail3=str(eMail3)

        #condizione eccesso idrico
        import smtplib
        #definiamo l'oggetto smptObj in cui si definisce il server smtp e la porta
        smtpObj= smtplib.SMTP('smtp.gmail.com', 587)
        #con il seguente comando diciamo "hello" al servizio smtp
        smtpObj.ehlo()
        #con il seguente comando definiamo lostandard di cifratura dei programmi TLS
        #dovrebbe andare bene per domini gMail. Potrebbe essere anche ssl (la porta sarebbe
        #465 invece che 587.in quel caso smtpObj=smtplib.SMTP_SSL('smtp.gmail.co', 465)
        #e il seguente comando sarebbe smtpObj.startssl().
        smtpObj.starttls()
        #mail: rpi.plant.out1@gmail.com
        #psw: quella normale è nA5!urzo; quella "per la app" è apiqwfffbctofcum
        #ora inseriamo i dati del mittente il mio indirizzo eMail e la password per le app che ho dovuto richiedere
        #a google previa verifica di sicurezza in due passaggi nell'impostazioni di
        #sicurezza dell' account google (password con invio codici via sms)  
        smtpObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
        #ATTENZIONE!!! Le credenziali sono giuste ma il server risponde Username and Password not accepted. Learn more at\n5.7.8 Capire perché
        #ora inseriamo rispettivamente: il mittente, il destinatario,
        #l'oggetto e il corpo della mail il corpo  separato dal soggetto dal carattere '\n'
        #l'importate per il comando \n (new line) che che non ci siano spazi dopo "\n" altrimenti non funziona 
        #oggetto=("c'è un eccesso idrico nel suolo, la bagnatura del suolo superficiale è ",bsm1," la bagnatura del suolo profonda è ", bsm2)

        #invio le eMail
        if(bsm1>LimDefIdr1 and bsm1<LimEccIdr1) and (bsm2>LimDefIdr2 and bsm2<LimEccIdr2):
            statoIdrico="acquaOk"
        else:
            statoIdrico="c'è un deficit o un eccesso di acqua: invio il messaggio"
            print (statoIdrico)
        if (eMail1!="None") and (statoIdrico!="acquaOk"):
            print("la eMail1 è stata impostata e  le invio un messaggio")
            import smtplib
            smtpObj= smtplib.SMTP('smtp.gmail.com', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
            smtpObj.sendmail('rpi.plant.out1@gmail.com', eMail1, "subject: "+oggetto)
        else:
            print("non ci sono le condizioni per l'invio del messaggio")
            
        if (eMail2!="None") and (statoIdrico!="acquaOk"):
            print("la eMail2 è stata impostata e  le invio un messaggio")
            import smtplib
            smtpObj= smtplib.SMTP('smtp.gmail.com', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
            smtpObj.sendmail('rpi.plant.out1@gmail.com', eMail2, "subject: "+oggetto)
        else:
            print("non ci sono le condizioni per l'invio del messaggio")       
        if (eMail3!="None") and (statoIdrico!="acquaOk"):
            print("la eMail3 è stata impostata e  le invio un messaggio")
            import smtplib
            smtpObj= smtplib.SMTP('smtp.gmail.com', 587)
            smtpObj.ehlo()
            smtpObj.starttls()
            smtpObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
            smtpObj.sendmail('rpi.plant.out1@gmail.com', eMail3, "subject: "+oggetto)
        else:
            print("non ci sono le condizioni per l'invio del messaggio")
        del smtplib    
        

    del openpyxl





#la prossima funzione controlla la posta e se c'è un messaggio con oggetto restoreGsheet ricostrisce gSheet
#da utilizzare in caso d'interruzione del collegamento Internet alla ripresa del collegamento
def restoreGsheet():
    #dalla versione 24 ripristina anche i dati orari
    import openpyxl        
    wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
    ws=wb["Sheet1"]
    ws1=wb["ultime48ore"]
    ws2=wb["ultimaSettimana"]
    ws3=wb["ultime2settimane"]
    ws4=wb["ultimoMese"]
    ws5=wb["datiGiornalieri"]
    ws8=wb["datiOrari"]
    from imapclient import IMAPClient
    server = IMAPClient('imap.gmail.com', use_uid=True)
    server.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')#credenziali di plantOut


    #così fa il login correttamente

    select_info = server.select_folder('INBOX')
    print('%d messages in INBOX' % select_info[b'EXISTS']) #mi dice i messaggi in inbox

    messages = server.search()
    print("%d Messaggi" % len(messages))#estrae il n°di messaggi

    for msgid, data in server.fetch(messages, ['ENVELOPE']).items():
        envelope = data[b'ENVELOPE']
        try:
            print('ID #%d: "%s" received %s' % (msgid, envelope.subject.decode(), envelope.date))#mi estrae n°del msg, oggetto, data.
        except(AttributeError):
            print('ID #%d: "[No Subject]" received %s' % (msgid, envelope.date))

    server.select_folder('INBOX', readonly=False)



    UIDs = server.search(['SUBJECT', 'restoreGsheet',]) #estrae i messaggi con soggetto restoreGsheet?
    print (UIDs)#mi mette le UIDs in una lista


    nmsg=len(UIDs)# il n ° di messaggi con oggetto restoreGsheet?

    print (nmsg)


    if nmsg!=0:
        print("c'è qualche messaggio con oggetto restoreGsheet, devo eseguire il codice  per ripristinare interamente RpiLogger")
        #inserire qui il codice per ripristinare gSheet
        daStoreAGsheet(ws,wks,8)
        print("e' stato ripristinato sheet1 su gStore")
        daStoreAGsheet(ws1,wks2,8)
        print("e' stato ripristinato ultime48ore su gStore")
        daStoreAGsheet(ws5,wks4,22)
        print("e' stato ripristinato datiGiornalieri su gStore")
        daStoreAGsheet(ws8,wks10,8)
        print("e' stato ripristinato datiOrari su gStore")
        daStoreAGsheet(ws2,wks5,22)
        print("e' stato ripristinato ultimaSettimana su gStore")
        daStoreAGsheet(ws3,wks6,22)
        print("e' stato ripristinato ultime2settimane su gStore")
        daStoreAGsheet(ws4,wks7,22)
        print("e' stato ripristinato ultimoMese su gStore")
        print("il restore di gSheet è stato completato")
        del openpyxl
        print (UIDs[0])
        print ("poi parte il codice che cancella il messaggio")
        import imapclient
        import imaplib
        imapObj=imapclient.IMAPClient("imap.gmail.com", 993)
        imapObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
        imapObj.select_folder('INBOX', readonly=False)
        imapObj.delete_messages(UIDs)
        imapObj.expunge()
        print("messaggio cancellato")
    else:
        print("non c'è nessun messaggio con oggetto restoreGsheet non parte il codice per ripristinare gSheet")
        
        
        
def restoreGsheetGiornaliero():
    #dalla versione 24 ripristina anche i datiOrari
    daStoreAGsheet(ws,wks,8)
    print("e' stato ripristinato sheet1 su gStore")
    daStoreAGsheet(ws1,wks2,8)
    print("e' stato ripristinato ultime48ore su gStore")
    daStoreAGsheet(ws5,wks4,22)
    print("e' stato ripristinato datiGiornalieri su gStore")
    daStoreAGsheet(ws8,wks10,8)
    print("e' stato ripristinato datiOrari su gStore")
    daStoreAGsheet(ws2,wks5,22)
    print("e' stato ripristinato ultimaSettimana su gStore")
    daStoreAGsheet(ws3,wks6,22)
    print("e' stato ripristinato ultime2settimane su gStore")
    daStoreAGsheet(ws4,wks7,22)
    print("e' stato ripristinato ultimoMese su gStore")

def rimandaVideoSuWeb():
    from imapclient import IMAPClient
    server = IMAPClient('imap.gmail.com', use_uid=True)
    server.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')#credenziali di plantOut


    #così fa il login correttamente

    select_info = server.select_folder('INBOX')
    print('%d messages in INBOX' % select_info[b'EXISTS']) #mi dice i messaggi in inbox

    messages = server.search()
    print("%d Messaggi" % len(messages))#estrae il n°di messaggi

    for msgid, data in server.fetch(messages, ['ENVELOPE']).items():
        envelope = data[b'ENVELOPE']
        try:
            print('ID #%d: "%s" received %s' % (msgid, envelope.subject.decode(), envelope.date))#mi estrae n°del msg, oggetto, data.
        except(AttributeError):
            print('ID #%d: "[No Subject]" received %s' % (msgid, envelope.date))

    server.select_folder('INBOX', readonly=False)



    UIDs = server.search(['SUBJECT', 'video',]) #estrae i messaggi con soggetto restoreGsheet?
    print (UIDs)#mi mette le UIDs in una lista


    nmsg=len(UIDs)# il n ° di messaggi con oggetto restoreGsheet?

    print (nmsg)


    if nmsg!=0:
        print("c'è qualche messaggio con oggetto video, devo eseguire il codice  per ripristinare interamente RpiLogger")
        mandaVideoSuWeb()
        print("ho inviato il video su web come richiesto mediante mail")
        print (UIDs[0])
        print ("poi parte il codice che cancella il messaggio")
        import imapclient
        import imaplib
        imapObj=imapclient.IMAPClient("imap.gmail.com", 993)
        imapObj.login('rpi.plant.out1@gmail.com', 'apiqwfffbctofcum')
        imapObj.select_folder('INBOX', readonly=False)
        imapObj.delete_messages(UIDs)
        imapObj.expunge()
        print("messaggio cancellato")
    else:
        print("non c'è nessun messaggio con oggetto video non parte il codice per invio de video su web")
        
def copiaERinominaStore():
    import datetime
    import os
    import shutil
    now=datetime.datetime.now()# chiedo che ora è
    ora=now
    print ("ora è proprio ...")
    print (ora)
    ora= ora.strftime('%d-%m-%Y_%He%M')
    print ("ora formattato nella maniera giusta e' "+ora)
    prefix="store"
    newFile=prefix+ora+".xslx"
    print ("il nuovo file sarà nominato "+newFile)
    src="/home/pi/plant+out/store.xlsx"
    dst="/media/pi/45D5-C1E61/copiaStore/store.xlsx"
    shutil.copy2(src, dst)
    print("ho copiato store nella directory copiaStore nella memoria usb")
    vecchioNome = "/media/pi/45D5-C1E61/copiaStore/store.xlsx"
    nuovoNome = "/media/pi/45D5-C1E61/copiaStore/"+newFile
    path1="/home/pi/plant+out/"+newFile
    path2="/media/pi/45D5-C1E61/copiaStore"
    os.rename(vecchioNome,nuovoNome)
    print("Ho rinominato store nella directory copiaStore nella memoria usb "+newFile)
    del datetime
    del os
    del shutil
    
def check_internet_connection():
    import requests
    try:
        # Prova a fare una richiesta a Google per verificare la connessione
        requests.get("https://www.google.com", timeout=5)
        return True
    except requests.ConnectionError:
        return False
    del requests        
    
#*******  Q U I   S O N O  F I N I T E  LE F U N Z I O N I  *********
   
    

#* * * * * * * * * * * *   R E C U P E R O   I   R I L I E V I   D A   S T O R E   * * * * * * * * * * * * * * * * * * * *


from datetime import datetime        

try:
    print("recupero i dati dei rilievi da store")
    #importo la libreria per scrivere su excell: 
    import openpyxl
    wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
    ws=wb["Sheet1"]
    #conta il numero di righe già scritte per scrivere i dati nella riga giusta
    nr=ws.max_row
    print ("la riga di lettura di ws di store è ",nr)

    taria=ws['B'+str(nr)].value
    print("l'ultimo rilievo di temperatura dell'aria è ",taria)

    RH=ws['C'+str(nr)].value
    print("l'ultimo rilievo di umidità relativa è ",RH)

    lux=ws['D'+str(nr)].value
    print("l'ultimo rilievo della luce è ",lux)


    bs1=ws['E'+str(nr)].value
    print("l'ultimo rilievo del sensore 1 di bagnatura del suolo in % in volume è  ",bs1)

    bs2=ws['G'+str(nr)].value
    print("l'ultimo rilievo del sensore 2 di bagnatura del suolo in % in volume è  ",bs2)
    print("deimporto la libreria openpyxl")
    del openpyxl
    ts1=20
    ts2=18
    print("la temperatura fittizia del suolo superficiale è ", ts1)
    print("la temperatura fittizia del suolo profonda è ", ts2)
    print("termino il programma")   
except:
    oraE=datetime.now()
    oraE=oraE.strftime("%H:%M:%S")
    print (oraE)
    filename=oraE+"errore_recuperoRilieviDaStore.txt"
    print(filename)
    #e lo inserisce in una specifica cartella "errori"
    errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
    errorFile.write(traceback.format_exc())
    errorFile.close()
    print('Le informazioni di traceback sono state scritte nel file ',filename) 







print("termino il programma")


"""
#*********DA QUI DOVREI FARE PARTIRE LE ALTRE 21 AZIONI IN ORDINE DI FREQUENZA************
#*********USO LE CONDIZIONI PER STABILIRE QUANDO DEVONO ESSERE LANCIATE*********************
# 1 F U N Z I O N E   C H E   P A R T E   C O N   I L   C R O N T A B   D U N Q U E   O G N I   1 5   M I N U T I

#***************** D E F I N I Z I O N I   D I   G O O G L E S H E E T   E   S T O R E ************
"""
if check_internet_connection():
    print("1 connessione attiva 1 copio i dati dei rilievi su gSheet")
    import gspread
    from oauth2client.service_account import ServiceAccountCredentials
    scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
    print (credentials)
    gc = gspread.authorize(credentials) 
    #nomino i fogli di GoogleSheet
    wks = gc.open("RpiPlantOut1Logger").sheet1


    scriviDatiSuGoogleFogli()#ogni 15 min. quindi senza temporizzatore
    del gspread
#scriviDatiSuGoogleFogli()#ogni 15 min. quindi senza temporizzatore
#print("dovrei avere eseguito scrividatiSuGoogleFogli")

#time.sleep(1)
#Da ora in poi le librerie per xl e gspread le deimporto alla line dello script   



#criteri per la versione del 26 marzo
#qui dovrei attaccare le altre funzioni senza controllare se la connessione è attiva
#a questo scopo indento le 21 funzioni perche parte delle condizione if
#e poi alla fine dare else dove attacco sole le funzioni che non implicano la connessione
#con la condizione dell'ora
    

    
      


    #definizioni di ora e minuti e giorno
    import openpyxl
    wb = wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
    ws=wb["Sheet1"]
    ultimaRiga=ws.max_row
    print(ultimaRiga)
    data=ws.cell(row=ultimaRiga,column=1).value
    print(data)
    print(type(data))
    ora=data[11]+data[12]
    print(ora)
    minu=data[14]+data[15]
    del openpyxl
    import datetime    
    now=datetime.datetime.now()# chiedo che ora è
    print (now)
    print (type(now))
    giorno=now.strftime('%A')
    print(giorno)
    del datetime

#2******** F O T O   C O N D I Z I O N E   L U X *********************
    try:
        print(minu)
        if minu=="00"or minu=="30":
            print("lancia 2 fotoCondizioneLux in 2 momenti corretti nell'ora")
            fotoCondizioneLux()
        else:
            print("non siamo ancora in 2 momenti corretti nell'ora che determinano il lancio di fotoCondizioneLux")    
    except BaseException as ex:
        import traceback
        import datetime
        from datetime import datetime
        oraE=datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreFotoCondizioneLux.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime

#3******** I N V I O   F O T O   S U   W E B *********************

    print(minu)
    if minu=="00"or minu=="30":
        print("lancia 3 invioFotoSuWeb in 2 momenti corretti nell'ora solo connessione se la connessione ok")
        try:
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione")
            invioFotoSuWeb()
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneInvioFotoSuWeb.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime

            
    else:
        print("connessione OK ma seconda condizione non soddisfatta, non lancio comunque invioFotosuWeb")


#4* * * * * * * *   C O P I A   U L T I M A   I N   F O T O   D I   P I A N T E   * * * * * * * * * * * * *
    try:
        import datetime
        if minu=="00"or minu=="30":
            print("lancia la 4 copiaUltimaInFotoDiPiante in 2 momenti corretti nell'ora")
            copiaUltimaInFotoDiPiante()
        else:
            print("non siamo ancora in 2 momenti corretti nell'ora che determinano il lancio di copiaUlimaInFotoDiPiante")    
    except BaseException as ex:
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreCopiaUltimaInFotoDiPiante.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime

#5 * * * * * * *   S C R I V I 4 8 O R E S T O R E   * * * * * * * * * * * * * * * * * *   
    try:
        print(minu)
        #openpyxl e wb  e ws già caricato all'interno di scrivi48oreStore
        if minu=="00":
            print("lancia la 5 48OreStore perché siamo al minuto dell'ora giusta")
            scrivi48oreStore()
            print("aggiorna anche i dati orari")
        else:
            print("non siamo ancora al minuto dell'ora giusta per lanciare la 48OreStore")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneScrivi48oreStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
    
#6 * * * * * * * * *   C O P I A   D A T I   4 8   O R E   S U   G S H E E T   * * * * * * *
    print(minu)
    if minu=="00":
        print("lancia la funzione 6 copia 48 ore su Gsheet perché siamo al minuto dell'ora giusta")
        try:
            #per la funzione daStoreAGsheet è necessario importare
            #openpyxl gspread wb, ws e wsb giusti prima di lanciare la funzione
            #infine deimportare entrambe le librerie
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione 6 copio foglio48ore su gSheet")
            import openpyxl
            wb = wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
            ws1=wb["ultime48ore"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks2 = gc.open("RpiPlantOut1Logger").worksheet('Ultime48Ore')
            daStoreAGsheet(ws1,wks2,8)
            del gspread
            del openpyxl
            
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneCopioFoglio48oreSugSheet.txt"
            print(filename)
             #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime

            
    else:
        print("connessione OK ma seconda condizione non soddisfatta, non lancio comunque la funzione copio dati 48 ore su gSheet")



    
#7 * * * * * * * * *   C O P I A   D A T I   O R A R I   S U   G S H E E T   * * * * * * *

    print(minu)
    if minu=="00":
        print("lancia la 7 copia dati orari su gSheet perché siamo al minuto dell'ora giusta")
        try:
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione copio dati orari su gSheet")
            #per la funzione daStoreAGsheet è necessario importare
            #openpyxl gspread wb, ws e wsb giusti prima di lanciare la funzione
            #infine deimportare entrambe le librerie
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione 7 copio dati orari su gSheet")
            import openpyxl
            wb = wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
            ws8=wb["datiOrari"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks10 = gc.open("RpiPlantOut1Logger").worksheet('datiOrari')
            daStoreAGsheet(ws8,wks10,8)
            del openpyxl
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneCopioFoglioDatiOrariSugSheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime          
    else:
        print("connessione OK ma condizione dell'ora non soddisfatta, non lancio comunque copio dati orari su gSheet")

    
#8 * *  C O N T R O L L A   M A I L   P E R   R E S T O R E       G S H E E T   * * *

    print(minu)
    if minu=="45":
        print("lancia la funzione perché siamo al minuto dell'ora giusta")
        try:
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione 8 restoregSheet")
            #non è necessario importare e nominare i fogli di store perchè contenuti già in restoreGsheet
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet inquanto non definiti in restoreGsheet
            wks = gc.open("RpiPlantOut1Logger").sheet1
            wks2 = gc.open("RpiPlantOut1Logger").worksheet('Ultime48Ore')
            wks4 = gc.open("RpiPlantOut1Logger").worksheet('DatiGiornalieri')
            wks5 = gc.open("RpiPlantOut1Logger").worksheet('UltimaSettimana')
            wks6 = gc.open("RpiPlantOut1Logger").worksheet('Ultime2Settimane')
            wks7 = gc.open("RpiPlantOut1Logger").worksheet('UltimoMese')
            wks9 = gc.open("RpiPlantOut1Logger").worksheet('Video')
            wks10 = gc.open("RpiPlantOut1Logger").worksheet('datiOrari')
            print("inserisco la funzione di controllo eMail per verifica se ricopiare store su gStore")
            restoreGsheet()
            del gspread
        except:
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneRestoreGsheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime

            
    else:
        print("connessione OK ma seconda condizione non soddisfatta, non lancio comunque restoreGsheet")

    


#9 * *  C O N T R O L L A   M A I L   P E R   R I N V I A R E   V I D E O   S U   W E B   * * *

    print(minu)
    if minu=="45":
        print("lancia la funzione rimandaVideoSuWeb perché siamo al minuto dell'ora giusta")
        try:
            #inserisci qui la funzione da lanciare
            print("inserisco la funzione di controllo eMail per verifica se rinviare il video su WEB")
            #la funzione usa l'indirizzo in locale depositato in store devo definirlo
            import openpyxl        
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws6=wb["videoInLocale"]#dst in colonna A; ora in colonna B
            rimandaVideoSuWeb()
            del openpyxl
        except:
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneRimandaVideoSuWeb.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime

            
    else:
        print("connessione OK ma seconda condizione non soddisfatta, non lancio comunque rimandaVideoSuWeb")

    
#10 * * * *   M E D I A   D A T I   G I O R N A L I E R I   * * * * * * * * 
    try: 
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione media dati giornalieri su store perche siamo all'ora e al minuto giusto")
            #non è necessario definire i fogli di store perché sono già definiti in mediaDatiGiornalieri
            mediaDatiGiornalieri()
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione media dati giornalieri")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneMediaDatiGiornalieriSuStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
    
#11 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M E   2   S E T T I M A N E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultime 2 settimane su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws3=wb["ultime2settimane"]
            scrivoDatiGiornalieriInAltriFogli(ws3,15)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultime 2 settimane store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInultime2settimane.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
    
#12 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M A  S E T T I M A N A   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultima settimana su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws2=wb["ultimaSettimana"]
            scrivoDatiGiornalieriInAltriFogli(ws2,8)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultima settimana store")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInUltimaSettimana.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
    
#13 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M O M E S E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultimo mese su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws4=wb["ultimoMese"]
            scrivoDatiGiornalieriInAltriFogli(ws4,32)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultimo mese store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInUltimoMese.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime

#14 * * * *   A G G I O R N A   D A T I   G I O R N A L I E R I   S U   G S H E E T   * * * * * * * *    

    print("14 connessione attiva per aggiorno dati giornalieri su gSheet")   
    print(ora)
    print(minu)
    if ora=="00"and minu=="15":
        try:
            print("lancia la funzione aggiorna dati giornalieri su gSheet perché siamo all'ora e al minuto giusto")
            #definisco i fogli gSheet perchÉ nella funzione aggiorna dati giornalieri su gSheet non sono definiti
            #non è necesssario definire i fogli di store perchè sono all'interno della funzione
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks4 = gc.open("RpiPlantOut1Logger").worksheet('DatiGiornalieri')
            aggiornaDatiGiornalieriSuGSheet()
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneAggiornaGiornalieriSuGsheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora all'ora e al minuto giusto per lanciare aggiornaGiornalieriSuGsheet")



#15 * * * *   C O P I A   U L T I M A   S E T T I M A N A   S U   G S H E E T    * * * * * * * * 
  
    print(ora)
    print(minu)
    if ora=="00"and minu=="15":
        try:
            print("lancia la funzione copia ultima settimana su gSheet perché siamo all'ora e al minuto giusto")
            #definisco i fogli di store
            #definisco i fogli gSheet perchÉ nella funzione daStoreAgsheet su gSheet non sono definiti
            import openpyxl        
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws2=wb["ultimaSettimana"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks5 = gc.open("RpiPlantOut1Logger").worksheet('UltimaSettimana')
            daStoreAGsheet(ws2,wks5,22)
            del openpyxl
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreCopiaUltimaSettimanaSuGsheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora all'ora e al minuto giusto per lanciare copiaUltimaSettimanaSuGsheet")

    

#16 * * * *   C O P I A   U L T I M E  2 S E T T I M A N E   S U   G S H E E T    * * * * * * * * 
  
    print(ora)
    print(minu)
    if ora=="00"and minu=="15":
        try:
            print("lancia la funzione copia ultime 2 settimane su gSheet perché siamo all'ora e al minuto giusto")
            #definisco i fogli di store
            #definisco i fogli gSheet perchÉ nella funzione daStoreAgsheet su gSheet non sono definiti
            import openpyxl        
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws3=wb["ultime2settimane"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks6 = gc.open("RpiPlantOut1Logger").worksheet('Ultime2Settimane')
            daStoreAGsheet(ws3,wks6,22)
            del openpyxl
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreCopiaUltime2settimaneSuGsheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora all'ora e al minuto giusto per lanciare copia ultime 2 settimane su gSheet")

    
#17 * * * *   C O P I A   U L T I M O M E S E  S U   G S H E E T    * * * * * * * * 
  
    print(ora)
    print(minu)
    if ora=="00"and minu=="15":
        try:
            print("lancia la funzione copia ultimo mese su gSheet perché siamo all'ora e al minuto giusto")
            #definisco i fogli di store
            #definisco i fogli gSheet perchÉ nella funzione daStoreAgsheet su gSheet non sono definiti
            import openpyxl        
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws4=wb["ultimoMese"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks7 = gc.open("RpiPlantOut1Logger").worksheet('UltimoMese')
            daStoreAGsheet(ws4,wks7,22)
            del openpyxl
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreCopiaUltimoMeseSuGsheet.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora all'ora e al minuto giusto per lanciare copia ultimo mese su gSheet")

    
#18 * * * *   F A   U N A   C O P I A   D I   S I C U R E Z Z A   D I   S T O R E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="16"and minu=="45":
            print("lancia la funzione che salva una copia di sicurezza di store su memoria USB perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            copiaERinominaStore()
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia di sicurezza di store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopiaERinominaStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime

        
#19 * * * *   M A I L   A L L A R M E   B A G N A T U R A   S U O L O    * * * * * * * *       

    print(ora)
    print(minu)
    if ora=="07"and minu=="15":
        try:
            print("lancia la funzione invio mail di allarme bagnatura suolo perché siamo all'ora e al minuto giusto")
            #nella funzione ci sono già le definizioni di store non è necessario definirle
            #ma la funzione usa anche i limiti inseriti in gSheet in wks13 devo definire gSheet e wks13
            #gSheet e non sono defiiti all'interno della funzione
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials)
            wks13 = gc.open("RpiPlantOut1Logger").worksheet('impostazioniPlant+out')
            sendAlarmBagnaturaSuolo()
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreSendAlarmBagnaturaSuolo.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora all'ora e al minuto giusto per lanciare sendAlarmBagnaturaSuolo")


#20 * * * *   C R E A   V I D E O    * * * * * * * *

    try:
        print(ora)
        print(minu)
        print(giorno)
        if (ora=="03"and minu=="45" and giorno=="Thursday") or (ora=="03"and minu=="45" and giorno=="Sunday"):
            print("lancia la funzione 20 crea videoperchÉ è il giorno, l'ora e il minuto giusto")
            #I fogli sono definiti all'interno della funzione, solo per scrivere la destinazione del video nel RPI 
            creaVideo()
        else:
            print("non siamo ancora al giorno, all'ora e al minuto giusto per lanciare la funzione crea video")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCreaVideo.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
        
#21 * * * *   I N V I A  V I D E O  S U   W E B     * * * * * * * *      

    print("21 connessione attiva per invio video su web")   
    print(ora)
    print(minu)
    print(giorno)
    if (ora=="04"and minu=="15" and giorno=="Thursday") or (ora=="04"and minu=="15" and giorno=="Sunday"):
        print("lancia la funzione 21 invia video su web perche' è il giorno, l'ora e il minuto giusto")
        try:
            print("lancia la funzione 21 invia video su web perche' è il giorno, l'ora e il minuto giusto")
            #definisco i fogli necessari di store e Gsheet perchÉ non sono definiti nella funzione
            import openpyxl        
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws6=wb["videoInLocale"]#dst in colonna A; ora in colonna B
            ws7=wb["videoSuWeb"]
            import gspread
            from oauth2client.service_account import ServiceAccountCredentials
            scope =  ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
            credentials = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/plant+out/plantout1logger-307322-7fff7cda5a13.json', scope)
            print (credentials)
            gc = gspread.authorize(credentials) 
            #nomino i fogli di GoogleSheet
            wks9 = gc.open("RpiPlantOut1Logger").worksheet('Video')
            mandaVideoSuWeb()
            del openpyxl          
            del gspread
        except:
            #in caso di errore nell'esecuzione del file, non perchè non c'è connessione
            import traceback
            import datetime
            oraE=datetime.datetime.now()
            oraE=oraE.strftime("%H:%M:%S")
            print (oraE)
            filename=oraE+"erroreEsecuzioneMandaVideoSuWeb.txt"
            print(filename)
            #e lo inserisce in una specifica cartella "errori"
            errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
            errorFile.write(traceback.format_exc())
            errorFile.close()
            print('Le informazioni di traceback sono state scritte nel file ',filename)
            del traceback
            del datetime  
    else:
        print("non siamo ancora al giorno, all'ora e al minuto giusto per lanciare mandaVideoSuWeb")

#***************** DA QUI IN POI VENGONO LANCIATE SOLO LE FUNZIONI CHE NON IMPLICANO CONNESSIONE*******
else: #attaccare qui tutte le funzioni che non implicano connessione
    print("non c'è connessione eseguo tutte le operazioni in locale")
    print("devo attaccare qui tutte le funzioni che partono se non c'è connessione")
    #definizioni di ora e minuti e giorno
    import openpyxl
    wb = wb=openpyxl.load_workbook("/home/pi/plant+out/store.xlsx", data_only=True)
    ws=wb["Sheet1"]
    ultimaRiga=ws.max_row
    print(ultimaRiga)
    data=ws.cell(row=ultimaRiga,column=1).value
    print(data)
    print(type(data))
    ora=data[11]+data[12]
    print(ora)
    minu=data[14]+data[15]
    del openpyxl
    import datetime    
    now=datetime.datetime.now()# chiedo che ora è
    print (now)
    print (type(now))
    giorno=now.strftime('%A')
    print(giorno)
    del datetime

    #2******** F O T O   C O N D I Z I O N E   L U X *********************
    try:
        print(minu)
        if minu=="00"or minu=="30":
            print("lancia 2 fotoCondizioneLux in 2 momenti corretti nell'ora")
            fotoCondizioneLux()
        else:
            print("non siamo ancora in 2 momenti corretti nell'ora che determinano il lancio di fotoCondizioneLux")    
    except BaseException as ex:
        import traceback
        import datetime
        from datetime import datetime
        oraE=datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreFotoCondizioneLux.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
      
    #4* * * * * * * *   C O P I A   U L T I M A   I N   F O T O   D I   P I A N T E   * * * * * * * * * * * * *
    try:
        import datetime
        if minu=="00"or minu=="30":
            print("lancia la 4 copiaUltimaInFotoDiPiante in 2 momenti corretti nell'ora")
            copiaUltimaInFotoDiPiante()
        else:
            print("non siamo ancora in 2 momenti corretti nell'ora che determinano il lancio di copiaUlimaInFotoDiPiante")    
    except BaseException as ex:
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreCopiaUltimaInFotoDiPiante.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime

    #5 * * * * * * *   S C R I V I 4 8 O R E S T O R E   * * * * * * * * * * * * * * * * * *   
    try:
        print(minu)
        #openpyxl e wb  e ws già caricato all'interno di scrivi48oreStore
        if minu=="00":
            print("lancia la 5 48OreStore perché siamo al minuto dell'ora giusta")
            scrivi48oreStore()
            print("aggiorna anche i dati orari")
        else:
            print("non siamo ancora al minuto dell'ora giusta per lanciare la 48OreStore")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneScrivi48oreStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
    #10 * * * *   M E D I A   D A T I   G I O R N A L I E R I   * * * * * * * * 
    try: 
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione media dati giornalieri su store perche siamo all'ora e al minuto giusto")
            #non è necessario definire i fogli di store perché sono già definiti in mediaDatiGiornalieri
            mediaDatiGiornalieri()
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione media dati giornalieri")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneMediaDatiGiornalieriSuStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
        
    #11 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M E   2   S E T T I M A N E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultime 2 settimane su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws3=wb["ultime2settimane"]
            scrivoDatiGiornalieriInAltriFogli(ws3,15)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultime 2 settimane store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInultime2settimane.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
        
    #12 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M A  S E T T I M A N A   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultima settimana su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws2=wb["ultimaSettimana"]
            scrivoDatiGiornalieriInAltriFogli(ws2,8)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultima settimana store")   
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInUltimaSettimana.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
        
    #13 * * * *   D A T I   G I O R N A L I E R I  I N  U L T I M O M E S E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="00"and minu=="00":
            print("lancia la funzione copio dati giornalieri in ultimo mese su store perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            import openpyxl
            wb=openpyxl.load_workbook('/home/pi/plant+out/store.xlsx',read_only=False)
            ws5=wb["datiGiornalieri"]
            ws4=wb["ultimoMese"]
            scrivoDatiGiornalieriInAltriFogli(ws4,32)
            del openpyxl
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia dati giornalieri in ultimo mese store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopioDatiGiornalieriInUltimoMese.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime    
       
    #18 * * * *   F A   U N A   C O P I A   D I   S I C U R E Z Z A   D I   S T O R E   * * * * * * * * 
    try:
        print(ora)
        print(minu)
        if ora=="16"and minu=="45":
            print("lancia la funzione che salva una copia di sicurezza di store su memoria USB perché siamo all'ora e al minuto giusto")
            #definisco i fogli perchÉ nella funzione non sono definiti
            copiaERinominaStore()
        else:
            print("non siamo ancora all'ora e al minuto giusto per lanciare la funzione copia di sicurezza di store")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCopiaERinominaStore.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
    #20 * * * *   C R E A   V I D E O    * * * * * * * *

    try:
        print(ora)
        print(minu)
        print(giorno)
        if (ora=="03"and minu=="45" and giorno=="Thursday") or (ora=="03"and minu=="45" and giorno=="Sunday"):
            print("lancia la funzione 20 crea videoperchÉ è il giorno, l'ora e il minuto giusto")
            #I fogli sono definiti all'interno della funzione, solo per scrivere la destinazione del video nel RPI 
            creaVideo()
        else:
            print("non siamo ancora al giorno, all'ora e al minuto giusto per lanciare la funzione crea video")    
    except:
        #in caso di errore nell'esecuzione del file, non perchè non siamo al momento giusto
        import traceback
        import datetime
        oraE=datetime.datetime.now()
        oraE=oraE.strftime("%H:%M:%S")
        print (oraE)
        filename=oraE+"erroreEsecuzioneFunzioneCreaVideo.txt"
        print(filename)
        #e lo inserisce in una specifica cartella "errori"
        errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
        errorFile.write(traceback.format_exc())
        errorFile.close()
        print('Le informazioni di traceback sono state scritte nel file ',filename)
        del traceback
        del datetime
    #
    import datetime
    from datetime import datetime
    oraE=datetime.now()
    oraE=oraE.strftime("%H:%M:%S")
    print (oraE)
    filename=oraE+"erroreConnessioneHoEseguitoIlPossibileInLocale.txt"
    print(filename)
    #e lo inserisce in una specifica cartella "errori"
    errorFile=open ('/home/pi/plant+out/errori/'+str(filename) ,'w')
    #errorFile.write()
    errorFile.close()
    print('Le informazioni di errore connessione sono state scritte nel file ',filename)
    del datetime
    
    
    
    
print("Termine del programma")